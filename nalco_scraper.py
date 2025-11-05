import os, sys, time, pathlib, re, requests, datetime
from urllib.parse import urljoin, urlparse
from bs4 import BeautifulSoup

import pdfplumber
import pandas as pd

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from zoneinfo import ZoneInfo

# ---------------------- CONFIG ----------------------
# Make all outputs relative to a base dir (defaults to current dir).
# This keeps behavior identical locally and on GitHub Actions.
BASE_DIR = pathlib.Path(os.getenv("BASE_DIR", ".")).resolve()

NALCO_URL = "https://nalcoindia.com/domestic/current-price/"
PDF_DIR = (BASE_DIR / "pdfs")
DATA_DIR = (BASE_DIR / "data")
LOG_FILE = DATA_DIR / "latest_nalco_pdf.txt"
EXCEL_FILE = DATA_DIR / "nalco_prices.xlsx"
RUNLOG_FILE = DATA_DIR / "nalco_run_log.xlsx"

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")

# prefer "Ingot-DD-MM-YYYY.pdf" and exclude spec docs
DATEY_PDF_RE = re.compile(r"Ingot-(\d{2})-(\d{2})-(\d{4})\.pdf$", re.IGNORECASE)

# Final Excel column order (DAILY layout)
DAILY_COLS = ["Date", "Description", "Product Code", "Basic Price", "Circular Date", "Circular Link"]

# Old layout columns (circular-wise rows) – we’ll accept these if present
OLD_COLS = ["Sl.no.", "Description", "Product Code", "Basic Price", "Circular Date", "Circular Link"]

# ---------------------- SCRAPER UTILS ----------------------
def ensure_dirs():
    for p in (PDF_DIR, DATA_DIR):
        if p.exists() and p.is_file():
            p.unlink()
        p.mkdir(parents=True, exist_ok=True)

def get_html(url):
    r = requests.get(url, headers={"User-Agent": UA}, timeout=60)
    r.raise_for_status()
    return r.text

def norm(s: str) -> str:
    return (s or "").strip().lower()

def find_ingots_pdf_url(html):
    soup = BeautifulSoup(html, "html.parser")

    # STRICT: <a ...><img ...><p>Ingots</p></a>
    for pnode in soup.find_all("p"):
        if norm(pnode.get_text()) == "ingots":
            a = pnode.find_parent("a", href=True)
            if a:
                href = a["href"].strip()
                if href.lower().endswith(".pdf") and "spec" not in href.lower():
                    return urljoin(NALCO_URL, href)

    # Prefer filenames like Ingot-DD-MM-YYYY.pdf (exclude spec)
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if href.lower().endswith(".pdf") and "spec" not in href.lower():
            if DATEY_PDF_RE.search(href):
                return urljoin(NALCO_URL, href)

    # Fallback: any PDF link whose anchor text mentions "ingot" (not spec)
    for a in soup.find_all("a", href=True):
        txt = norm(a.get_text())
        href = a["href"].strip()
        if href.lower().endswith(".pdf") and "ingot" in txt and "spec" not in href.lower():
            return urljoin(NALCO_URL, href)

    return None

def read_last_url():
    return LOG_FILE.read_text(encoding="utf-8").strip() if LOG_FILE.exists() else ""

def write_last_url(url):
    LOG_FILE.write_text(url or "", encoding="utf-8")

def download_pdf(url):
    headers = {
        "User-Agent": UA,
        "Referer": NALCO_URL,
        "Accept": "application/pdf,*/*;q=0.9",
        "Accept-Language": "en-US,en;q=0.9",
    }
    with requests.get(url, headers=headers, timeout=60, stream=True, allow_redirects=True) as r:
        r.raise_for_status()
        ctype = r.headers.get("Content-Type", "").lower()
        if "application/pdf" not in ctype:
            raise RuntimeError(f"Expected PDF but got Content-Type={ctype!r} from {url}")
        filename = None
        cd = r.headers.get("Content-Disposition", "")
        if "filename=" in cd:
            filename = cd.split("filename=", 1)[1].strip('"; ')
        if not filename:
            filename = os.path.basename(urlparse(r.url).path) or f"nalco_{int(time.time())}.pdf"
        dest = PDF_DIR / filename
        with open(dest, "wb") as f:
            for chunk in r.iter_content(chunk_size=65536):
                if chunk:
                    f.write(chunk)
        return dest

# ---------------------- PDF PARSING ----------------------
def parse_circular_date_from_filename(pdf_path: pathlib.Path) -> str:
    """Extract DD-MM-YYYY from 'Ingot-DD-MM-YYYY.pdf', else use today."""
    m = DATEY_PDF_RE.search(pdf_path.name)
    if m:
        dd, mm, yyyy = m.groups()
        try:
            d = datetime.date(int(yyyy), int(mm), int(dd))
            return d.strftime("%d-%m-%Y")
        except ValueError:
            pass
    return datetime.date.today().strftime("%d-%m-%Y")

def extract_row_ie07(pdf_path: pathlib.Path):
    """
    Find the row that contains 'IE07'.
    Returns dict: { 'Description', 'Product Code', 'Basic Price' }
    """
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            # Try structured tables
            try:
                tables = page.extract_tables()
            except Exception:
                tables = []
            for tbl in tables or []:
                for row in tbl:
                    cells = [(c or "").strip() for c in row]
                    if any(re.fullmatch(r"IE07", x, flags=re.IGNORECASE) for x in cells):
                        desc = ""
                        code = "IE07"
                        price = ""
                        idx_code = None
                        for i, c in enumerate(cells):
                            if re.fullmatch(r"IE07", c, flags=re.IGNORECASE):
                                idx_code = i
                                break
                        if idx_code is not None:
                            # description: look left for non-numeric text
                            for j in range(idx_code - 1, -1, -1):
                                if cells[j] and not re.fullmatch(r"\d+(\.\d+)?", cells[j]):
                                    desc = cells[j]
                                    break
                            # price: first numeric-looking to right
                            for j in range(idx_code + 1, len(cells)):
                                if re.search(r"\d", cells[j]):
                                    price = cells[j].replace(",", "")
                                    break
                        if code and price:
                            return {
                                "Description": (desc or "ALUMINIUM INGOT").upper(),
                                "Product Code": code,
                                "Basic Price": price
                            }

            # Fallback: line text scan
            words = page.extract_words(use_text_flow=True, keep_blank_chars=False)
            lines = {}
            for w in words:
                y = round(w["top"], 1)
                lines.setdefault(y, []).append(w)
            for y, wlist in lines.items():
                text_line = " ".join([w["text"] for w in sorted(wlist, key=lambda x: x["x0"])])
                if re.search(r"\bIE07\b", text_line, flags=re.IGNORECASE):
                    m_price = re.search(r"(\d{5,7}(?:\.\d+)?)\s*$", text_line)
                    price = (m_price.group(1) if m_price else "").replace(",", "")
                    m_desc = re.search(r"([A-Z ]*INGOT[A-Z ]*)\b", text_line, flags=re.IGNORECASE)
                    desc = (m_desc.group(1) if m_desc else "ALUMINIUM INGOT").strip().upper()
                    if price:
                        return {
                            "Description": desc,
                            "Product Code": "IE07",
                            "Basic Price": price
                        }
    raise RuntimeError("Could not find a row with Product Code IE07 in the PDF.")

def to_thousands(value_str: str) -> float:
    """Convert raw price (e.g., '268250') to thousands (e.g., 268.250)."""
    value_str = value_str.replace(",", "").strip()
    if not value_str:
        return None
    try:
        v = float(value_str)
        return round(v / 1000.0, 3)
    except ValueError:
        return None

# ---------------------- EXCEL HELPERS ----------------------
def save_excel_formatted(df: pd.DataFrame, path: pathlib.Path):
    """
    Save df to Excel, auto-fit column widths based on content,
    center-align all cells, and make Circular Link clickable.
    """
    df.to_excel(path, index=False)

    wb = load_workbook(path)
    ws = wb.active

    center = Alignment(horizontal="center", vertical="center")

    # Auto width
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = len(str(col_name))
        for val in df[col_name].astype(str).values:
            if val is None:
                continue
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_len + 2, 80))

    header_row = 1
    nrows = ws.max_row
    ncols = ws.max_column
    # indices in DAILY layout
    price_col_idx = DAILY_COLS.index("Basic Price") + 1
    link_col_idx  = DAILY_COLS.index("Circular Link") + 1

    for r in range(1, nrows + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = center
            if r > header_row and c == price_col_idx:
                cell.number_format = "0.000"
            if r > header_row and c == link_col_idx:
                val = cell.value
                if isinstance(val, str) and val.startswith("http"):
                    cell.hyperlink = val

    ws.freeze_panes = "A2"
    wb.save(path)

def sort_for_display_daily(df: pd.DataFrame) -> pd.DataFrame:
    """Sort by Date descending for display. Keep formats consistent."""
    dtd = pd.to_datetime(df["Date"], dayfirst=True, errors="coerce")
    df = df.assign(_date=dtd)
    df = df.sort_values(by=["_date"], ascending=[False], kind="stable").drop(columns=["_date"])
    df["Basic Price"] = pd.to_numeric(df["Basic Price"], errors="coerce").round(3)
    df["Date"] = pd.to_datetime(df["Date"], dayfirst=True, errors="coerce").dt.strftime("%d-%m-%Y")
    df["Circular Date"] = pd.to_datetime(df["Circular Date"], dayfirst=True, errors="coerce").dt.strftime("%d-%m-%Y")
    return df

# ---------------------- DAILY SERIES BUILDER ----------------------
def build_daily_df_from_circulars(circ_df: pd.DataFrame) -> pd.DataFrame:
    """
    Given a dataframe of circular records (one per circular), return a DAILY dataframe:
      Columns: DAILY_COLS
      Rows: every calendar date from first circular date to YESTERDAY (IST)
      Values: forward-fill the last known circular's price/link/date.
    """
    if circ_df.empty:
        return pd.DataFrame(columns=DAILY_COLS)

    # Ensure types
    circ_df = circ_df.copy()
    circ_df["Circular Date"] = pd.to_datetime(circ_df["Circular Date"], dayfirst=True, errors="coerce")
    circ_df["Basic Price"] = pd.to_numeric(circ_df["Basic Price"], errors="coerce").round(3)
    circ_df = circ_df.sort_values("Circular Date")

    # Index by circular date
    circ_df_idx = circ_df.set_index("Circular Date")[["Description", "Product Code", "Basic Price", "Circular Link"]]

    # Build full daily range (to yesterday IST)
    yesterday_ist = datetime.datetime.now(ZoneInfo("Asia/Kolkata")).date() - datetime.timedelta(days=1)
    start_date = circ_df["Circular Date"].min().date()
    if start_date > yesterday_ist:
        # If first circular is in the future (unlikely), return empty
        return pd.DataFrame(columns=DAILY_COLS)

    full_range = pd.date_range(start=start_date, end=yesterday_ist, freq="D")

    # Reindex and forward-fill
    filled = circ_df_idx.reindex(full_range).ffill()
    filled.index.name = "Date"
    daily = filled.reset_index()

    # We also want the "Circular Date" column to reflect the last circular date that applied:
    # After ffill, the "where it came from" is the index of last valid entry. We can compute
    # by merging against the forward-filled mask of original dates.
    # Easiest: carry a helper series of the original index and ffill it too.
    key = pd.Series(circ_df.set_index("Circular Date").index, index=circ_df.set_index("Circular Date").index, name="Circular Date").reindex(full_range).ffill()
    daily["Circular Date"] = key.values

    # Final ordering & formatting
    daily = daily.rename(columns={"Date": "Date"})
    daily["Date"] = pd.to_datetime(daily["Date"]).dt.strftime("%d-%m-%Y")
    daily["Circular Date"] = pd.to_datetime(daily["Circular Date"]).dt.strftime("%d-%m-%Y")
    daily = daily[DAILY_COLS]

    # Display sort (desc by Date)
    daily = sort_for_display_daily(daily)
    return daily

def derive_circulars_from_existing(df: pd.DataFrame) -> pd.DataFrame:
    """
    Accept either old layout (Sl.no.-based) or daily layout and return a *circulars-only* dataframe:
      columns: Description, Product Code, Basic Price, Circular Date, Circular Link
      one row per unique Circular Date (last one wins if duplicates)
    """
    if df.empty:
        return pd.DataFrame(columns=["Description", "Product Code", "Basic Price", "Circular Date", "Circular Link"])

    temp = df.copy()

    # Normalize column names from either old or new layout
    cols = {c.lower(): c for c in temp.columns}
    # Required logical columns (case-insensitive lookup)
    def get(colname):
        # try exact, then case-insensitive
        return cols.get(colname) or next((c for c in temp.columns if c.strip().lower() == colname), None)

    # Map columns
    col_desc = get("description")
    col_code = get("product code")
    col_price= get("basic price")
    col_cdate= get("circular date")
    col_link = get("circular link")
    if not all([col_desc, col_code, col_price, col_cdate, col_link]):
        # Unknown layout, return empty (we'll rebuild from only new circular if any)
        return pd.DataFrame(columns=["Description", "Product Code", "Basic Price", "Circular Date", "Circular Link"])

    circ = temp[[col_desc, col_code, col_price, col_cdate, col_link]].copy()
    circ.columns = ["Description", "Product Code", "Basic Price", "Circular Date", "Circular Link"]

    # Reduce to one row per Circular Date (keep the latest entry per date)
    circ["Circular Date"] = pd.to_datetime(circ["Circular Date"], dayfirst=True, errors="coerce")
    circ = circ.dropna(subset=["Circular Date"])
    circ = circ.sort_values("Circular Date").drop_duplicates(subset=["Circular Date"], keep="last")

    # Ensure numeric price
    circ["Basic Price"] = pd.to_numeric(circ["Basic Price"], errors="coerce").round(3)
    return circ

# ---------------------- RUN LOG HELPERS ----------------------
def append_runlog(log_path: pathlib.Path, info: dict):
    """
    Append (or create) a run log Excel separate from the main data file.
    Columns: Run UTC, Run IST, Status, Message, Chosen URL, Saved PDF, Rows Appended, Total Rows After
    """
    cols = ["Run UTC", "Run IST", "Status", "Message", "Chosen URL", "Saved PDF", "Rows Appended", "Total Rows After"]
    if log_path.exists():
        df = pd.read_excel(log_path)
        for c in cols:
            if c not in df.columns:
                df[c] = pd.NA
        df = df[cols]
    else:
        df = pd.DataFrame(columns=cols)

    df = pd.concat([df, pd.DataFrame([{
        "Run UTC": info.get("Run UTC"),
        "Run IST": info.get("Run IST"),
        "Status": info.get("Status"),
        "Message": info.get("Message"),
        "Chosen URL": info.get("Chosen URL"),
        "Saved PDF": info.get("Saved PDF"),
        "Rows Appended": info.get("Rows Appended"),
        "Total Rows After": info.get("Total Rows After"),
    }])], ignore_index=True)

    # Save & simple formatting (autofit + center)
    df.to_excel(log_path, index=False)
    wb = load_workbook(log_path)
    ws = wb.active
    center = Alignment(horizontal="center", vertical="center")
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = len(str(col_name))
        for val in df[col_name].astype(str).values:
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_len + 2, 100))
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).alignment = center
    ws.freeze_panes = "A2"
    wb.save(log_path)

def now_times():
    now_utc = datetime.datetime.now(datetime.timezone.utc)
    now_ist = now_utc.astimezone(ZoneInfo("Asia/Kolkata"))
    return now_utc.strftime("%Y-%m-%d %H:%M:%S UTC"), now_ist.strftime("%Y-%m-%d %H:%M:%S IST")

# ---------------------- MAIN FLOW ----------------------
def main():
    ensure_dirs()
    run_utc, run_ist = now_times()

    # Step 1: scrape current page for latest Ingot PDF
    html = get_html(NALCO_URL)
    pdf_url = find_ingots_pdf_url(html)
    chosen_pdf_path = ""
    rows_appended = 0

    if not pdf_url:
        msg = "No Ingots PDF link found on the page."
        print(msg, file=sys.stderr)
        # even if no new URL, we still rebuild daily from what's already in Excel
        existing_df = pd.read_excel(EXCEL_FILE) if EXCEL_FILE.exists() else pd.DataFrame()
        circ_df = derive_circulars_from_existing(existing_df)
        daily_df = build_daily_df_from_circulars(circ_df)
        if not daily_df.empty:
            save_excel_formatted(daily_df[DAILY_COLS], EXCEL_FILE)
            print(f"Excel rebuilt (daily): {EXCEL_FILE}")
        append_runlog(RUNLOG_FILE, {
            "Run UTC": run_utc, "Run IST": run_ist, "Status": "SKIPPED",
            "Message": msg, "Chosen URL": "", "Saved PDF": "", "Rows Appended": rows_appended,
            "Total Rows After": daily_df.shape[0] if not daily_df.empty else 0
        })
        sys.exit(0)

    print(f"Chosen PDF URL: {pdf_url}")

    last = read_last_url()
    new_circular_detected = (pdf_url != last)

    if new_circular_detected:
        # Step 2: fetch and parse new circular
        pdf_path = download_pdf(pdf_url)
        write_last_url(pdf_url)
        chosen_pdf_path = pdf_path.name
        print(f"Saved to: {pdf_path}")

        # Extract IE07 row from PDF
        row = extract_row_ie07(pdf_path)

        # Convert Basic Price to thousands (e.g., 268250 -> 268.250)
        thousands = to_thousands(row["Basic Price"])
        if thousands is None:
            raise RuntimeError(f"Could not parse numeric price from: {row['Basic Price']!r}")
        row["Basic Price"] = thousands

        # Circular date: from filename (fallback to today)
        row["Circular Date"] = parse_circular_date_from_filename(pdf_path)
        # Link: the PDF URL used
        row["Circular Link"] = pdf_url
        row["Description"] = row.get("Description", "ALUMINIUM INGOT")
        row["Product Code"] = row.get("Product Code", "IE07")

        # Build circulars dataframe from existing data, then add this circular (by Circular Date)
        existing_df = pd.read_excel(EXCEL_FILE) if EXCEL_FILE.exists() else pd.DataFrame()
        circ_df = derive_circulars_from_existing(existing_df)

        # Upsert this circular by its Circular Date
        new_c = pd.DataFrame([{
            "Description": row["Description"],
            "Product Code": row["Product Code"],
            "Basic Price": row["Basic Price"],
            "Circular Date": pd.to_datetime(row["Circular Date"], dayfirst=True, errors="coerce"),
            "Circular Link": row["Circular Link"],
        }])
        circ_df = pd.concat([circ_df, new_c], ignore_index=True)
        circ_df = circ_df.dropna(subset=["Circular Date"]).sort_values("Circular Date").drop_duplicates(subset=["Circular Date"], keep="last")

        # Step 3: rebuild DAILY sheet to yesterday IST
        daily_df = build_daily_df_from_circulars(circ_df)
        save_excel_formatted(daily_df[DAILY_COLS], EXCEL_FILE)
        rows_appended = 1  # one new circular processed

        append_runlog(RUNLOG_FILE, {
            "Run UTC": run_utc, "Run IST": run_ist, "Status": "UPDATED",
            "Message": "New circular processed and daily sheet rebuilt.",
            "Chosen URL": pdf_url, "Saved PDF": chosen_pdf_path,
            "Rows Appended": rows_appended, "Total Rows After": daily_df.shape[0]
        })
        return

    else:
        # No new PDF since last run – still rebuild DAILY sheet up to yesterday IST.
        msg = "No change in PDF. Rebuilding daily sheet up to yesterday."
        print(msg)
        existing_df = pd.read_excel(EXCEL_FILE) if EXCEL_FILE.exists() else pd.DataFrame()
        circ_df = derive_circulars_from_existing(existing_df)
        daily_df = build_daily_df_from_circulars(circ_df)
        if not daily_df.empty:
            save_excel_formatted(daily_df[DAILY_COLS], EXCEL_FILE)
            print(f"Excel rebuilt (daily): {EXCEL_FILE}")

        append_runlog(RUNLOG_FILE, {
            "Run UTC": run_utc, "Run IST": run_ist, "Status": "SKIPPED",
            "Message": msg, "Chosen URL": pdf_url, "Saved PDF": "",
            "Rows Appended": rows_appended, "Total Rows After": daily_df.shape[0] if not daily_df.empty else 0
        })
        return

if __name__ == "__main__":
    main()
